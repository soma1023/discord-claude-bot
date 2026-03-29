import discord
import subprocess
import os
import sys
import asyncio
import base64
import aiohttp
import anthropic
import openpyxl
import send2trash
import json
import time
from datetime import datetime
from dotenv import load_dotenv

# ローカル .env（相手ユーザーのPC用）を優先して読み込む
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), '.env'))
# オーナーPC用のパスからも読み込む
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), '..', '.claude', 'channels', 'discord', '.env'))
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), '..', 'discord-job-manager', '.env'))

TOKEN = os.getenv('DISCORD_BOT_TOKEN')
ANTHROPIC_API_KEY = os.getenv('ANTHROPIC_API_KEY')

# .envから設定を読み込む（各PCで独自に設定）
ALLOWED_USER_IDS = {
    int(uid.strip())
    for uid in os.getenv('ALLOWED_USER_IDS', '').split(',')
    if uid.strip()
}
WORK_DIR = os.path.expanduser('~')
MAX_HISTORY = 20  # 1ユーザーあたりの最大メッセージ数（往復10回）
LOG_PATH = os.path.join(os.path.dirname(__file__), 'logs', 'usage.jsonl')
os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)

SYSTEM_PROMPT = """あなたはClaudeです。ユーザーのDiscordから指示を受け、PCを操作するAIアシスタントです。
ファイルの読み書き、コード実行、コマンド実行など、Claude Codeと同等の作業ができます。
作業内容は日本語で簡潔に報告してください。
OSはWindows 11です。コマンドはbashまたはPowerShell構文で実行できます。"""

TOOLS = [
    {
        "name": "bash",
        "description": "シェルコマンドを実行する。Windows環境なのでbash/PowerShell両方使える。",
        "input_schema": {
            "type": "object",
            "properties": {
                "command": {"type": "string", "description": "実行するコマンド"}
            },
            "required": ["command"]
        }
    },
    {
        "name": "read_file",
        "description": "テキストファイルの内容を読み込む。",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "読み込むファイルのパス"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "write_file",
        "description": "ファイルに内容を書き込む（上書き）。",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "書き込むファイルのパス"},
                "content": {"type": "string", "description": "書き込む内容"}
            },
            "required": ["path", "content"]
        }
    },
    {
        "name": "send_file",
        "description": "PCにあるファイル（画像・テキスト・コードなど）をDiscordチャットに送信する。",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "送信するファイルのパス"},
                "caption": {"type": "string", "description": "ファイルに添えるコメント（省略可）"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "read_image",
        "description": "PC上の画像ファイルを読み込んで内容を解析する。スクリーンショット・図・写真など。",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "解析する画像ファイルのパス（png/jpg/gif/webp）"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "list_files",
        "description": "ディレクトリのファイル一覧を取得する。",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "一覧を取得するディレクトリのパス（省略時はホームディレクトリ）"}
            },
            "required": []
        }
    },
    {
        "name": "search_files",
        "description": "ファイル内のテキストを検索する（grep相当）。",
        "input_schema": {
            "type": "object",
            "properties": {
                "pattern": {"type": "string", "description": "検索パターン（正規表現可）"},
                "path": {"type": "string", "description": "検索対象のパスまたはディレクトリ"}
            },
            "required": ["pattern", "path"]
        }
    },
    {
        "name": "trash",
        "description": "ファイルまたはフォルダをゴミ箱に移動する（完全削除ではなく復元可能）。",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "ゴミ箱に移動するファイル/フォルダのパス"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "excel_read",
        "description": "Excelファイルのセル範囲を読み込む。シート名省略時はアクティブシートを使用。",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Excelファイルのパス（.xlsx）"},
                "sheet": {"type": "string", "description": "シート名（省略時はアクティブシート）"},
                "range": {"type": "string", "description": "セル範囲（例: A1:C10）。省略時はデータ全体"}
            },
            "required": ["path"]
        }
    },
    {
        "name": "excel_write",
        "description": "Excelファイルの指定セルに値を書き込む。ファイルが存在しない場合は新規作成。",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Excelファイルのパス（.xlsx）"},
                "sheet": {"type": "string", "description": "シート名（省略時はアクティブシート）"},
                "cell": {"type": "string", "description": "セル番地（例: B3）"},
                "value": {"description": "書き込む値（文字列・数値・日付など）"}
            },
            "required": ["path", "cell", "value"]
        }
    },
    {
        "name": "excel_append",
        "description": "Excelファイルの最終行の次に新しい行を追加する。",
        "input_schema": {
            "type": "object",
            "properties": {
                "path": {"type": "string", "description": "Excelファイルのパス（.xlsx）"},
                "sheet": {"type": "string", "description": "シート名（省略時はアクティブシート）"},
                "values": {
                    "type": "array",
                    "description": "追加する行のデータ（例: [\"山田\", 25, \"東京\"]）",
                    "items": {}
                }
            },
            "required": ["path", "values"]
        }
    }
]

intents = discord.Intents.default()
intents.message_content = True
client = discord.Client(intents=intents)
ai = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

# ユーザーごとの会話履歴 { user_id: [{role, content}, ...] }
history: dict[int, list] = {}


def log_event(entry: dict):
    """操作ログをJSONL形式で追記する"""
    try:
        entry["ts"] = datetime.now().isoformat(timespec='seconds')
        with open(LOG_PATH, 'a', encoding='utf-8') as f:
            f.write(json.dumps(entry, ensure_ascii=False) + '\n')
    except Exception:
        pass


def get_history(user_id: int) -> list:
    return history.setdefault(user_id, [])


def add_to_history(user_id: int, role: str, content: str):
    h = get_history(user_id)
    h.append({"role": role, "content": content})
    while len(h) > MAX_HISTORY:
        h.pop(0)
        h.pop(0)


def execute_tool(name: str, inp: dict, user_id: int = 0) -> str:
    out = None
    success = True
    try:
        if name == "bash":
            proc = subprocess.run(
                inp["command"], shell=True, capture_output=True,
                text=True, encoding='utf-8', errors='replace',
                timeout=60, cwd=WORK_DIR
            )
            raw = (proc.stdout + proc.stderr).strip()
            out = raw[:8000] if raw else "（出力なし）"

        elif name == "read_file":
            path = os.path.expanduser(inp["path"])
            with open(path, 'r', encoding='utf-8', errors='replace') as f:
                out = f.read()[:8000]

        elif name == "write_file":
            path = os.path.expanduser(inp["path"])
            os.makedirs(os.path.dirname(path), exist_ok=True) if os.path.dirname(path) else None
            with open(path, 'w', encoding='utf-8') as f:
                f.write(inp["content"])
            out = f"書き込み完了: {path}"

        elif name == "list_files":
            path = os.path.expanduser(inp.get("path", WORK_DIR))
            entries = os.listdir(path)
            lines = []
            for e in sorted(entries):
                full = os.path.join(path, e)
                tag = "/" if os.path.isdir(full) else ""
                lines.append(f"{e}{tag}")
            out = "\n".join(lines) or "（空）"

        elif name == "search_files":
            proc = subprocess.run(
                ["grep", "-r", "-n", inp["pattern"], os.path.expanduser(inp["path"])],
                capture_output=True, text=True, encoding='utf-8',
                errors='replace', timeout=30
            )
            raw = proc.stdout.strip()
            out = raw[:8000] if raw else "マッチなし"

        elif name == "trash":
            path = os.path.expanduser(inp["path"])
            send2trash.send2trash(path)
            out = f"ゴミ箱に移動しました: {path}"

        elif name == "excel_read":
            path = os.path.expanduser(inp["path"])
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[inp["sheet"]] if inp.get("sheet") else wb.active
            if inp.get("range"):
                cells = ws[inp["range"]]
                if not isinstance(cells, (tuple, list)):
                    rows = [[cells.value]]
                elif cells and not isinstance(cells[0], (tuple, list)):
                    rows = [[c.value for c in cells]]
                else:
                    rows = [[c.value for c in row] for row in cells]
            else:
                rows = [[c.value for c in row] for row in ws.iter_rows()]
            lines = []
            for row in rows:
                lines.append("\t".join("" if v is None else str(v) for v in row))
            out = "\n".join(lines)[:8000] or "（データなし）"

        elif name == "excel_write":
            path = os.path.expanduser(inp["path"])
            wb = openpyxl.load_workbook(path) if os.path.exists(path) else openpyxl.Workbook()
            ws = wb[inp["sheet"]] if inp.get("sheet") and inp["sheet"] in wb.sheetnames else wb.active
            ws[inp["cell"]] = inp["value"]
            wb.save(path)
            out = f"{inp['cell']} に '{inp['value']}' を書き込みました: {path}"

        elif name == "excel_append":
            path = os.path.expanduser(inp["path"])
            wb = openpyxl.load_workbook(path) if os.path.exists(path) else openpyxl.Workbook()
            ws = wb[inp["sheet"]] if inp.get("sheet") and inp["sheet"] in wb.sheetnames else wb.active
            ws.append(inp["values"])
            wb.save(path)
            out = f"行を追加しました（{len(inp['values'])}列）: {path}"

    except subprocess.TimeoutExpired:
        out = "タイムアウト（60秒）"
        success = False
    except Exception as e:
        out = f"エラー: {e}"
        success = False

    log_event({
        "event": "tool_use",
        "user_id": user_id,
        "tool": name,
        "input": inp,
        "result_preview": (out or "")[:300],
        "success": success
    })
    return out


async def update_status_from_headers(headers):
    """レート制限ヘッダーからDiscordステータスを更新する"""
    try:
        remaining = headers.get('anthropic-ratelimit-tokens-remaining')
        limit = headers.get('anthropic-ratelimit-tokens-limit')
        if remaining and limit:
            pct = int(remaining) / int(limit)
            if pct > 0.6:
                status = discord.Status.online
            elif pct > 0.3:
                status = discord.Status.idle
            else:
                status = discord.Status.dnd
            await client.change_presence(status=status)
    except Exception:
        pass


async def run_agent(user_id: int, user_content, channel) -> str:
    """ツール使用を含むエージェントループ"""
    # 履歴＋今回のメッセージでmessagesを構築
    messages = list(get_history(user_id)) + [{"role": "user", "content": user_content}]

    tool_count = 0
    while True:
        # 429レート制限時に指数バックオフでリトライ（最大3回）
        for attempt in range(4):
            try:
                raw = await asyncio.get_event_loop().run_in_executor(
                    None,
                    lambda msgs=messages: ai.messages.with_raw_response.create(
                        model="claude-sonnet-4-6",
                        max_tokens=4096,
                        system=SYSTEM_PROMPT,
                        tools=TOOLS,
                        messages=msgs
                    )
                )
                break
            except anthropic.RateLimitError:
                if attempt == 3:
                    raise
                wait = 2 ** attempt * 5  # 5秒, 10秒, 20秒
                await channel.send(f"⏳ レート制限中。{wait}秒後にリトライします... ({attempt + 1}/3)")
                await asyncio.sleep(wait)
        response = raw.parse()
        await update_status_from_headers(raw.headers)

        if response.stop_reason != "tool_use":
            # 最終回答
            final = next(
                (b.text for b in response.content if hasattr(b, 'text')),
                "（応答なし）"
            )
            # 履歴に追加（ユーザーメッセージ＋最終回答のみ）
            user_text = user_content if isinstance(user_content, str) else "[画像+テキスト]"
            add_to_history(user_id, "user", user_text)
            add_to_history(user_id, "assistant", final)
            log_event({
                "event": "message",
                "user_id": user_id,
                "user_text": user_text[:300],
                "response_preview": final[:300],
                "tool_count": tool_count,
                "input_tokens": getattr(response.usage, "input_tokens", None),
                "output_tokens": getattr(response.usage, "output_tokens", None)
            })
            return final

        # ツール実行
        messages.append({"role": "assistant", "content": response.content})
        tool_results = []

        for block in response.content:
            if block.type != "tool_use":
                continue

            tool_count += 1

            # send_file: 非同期でDiscordにファイルを送信
            if block.name == "send_file":
                path = os.path.expanduser(block.input["path"])
                caption = block.input.get("caption", "")
                try:
                    await channel.send(
                        content=caption if caption else None,
                        file=discord.File(path)
                    )
                    result_content = f"ファイルを送信しました: {path}"
                    log_event({"event": "tool_use", "user_id": user_id, "tool": "send_file",
                               "input": block.input, "result_preview": result_content, "success": True})
                except Exception as e:
                    result_content = f"送信エラー: {e}"
                    log_event({"event": "tool_use", "user_id": user_id, "tool": "send_file",
                               "input": block.input, "result_preview": result_content, "success": False})
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": block.id,
                    "content": result_content
                })

            # read_image: 画像をbase64で読み込んでClaudeに渡す
            elif block.name == "read_image":
                path = os.path.expanduser(block.input["path"])
                try:
                    ext = os.path.splitext(path)[1].lower()
                    media_map = {".png": "image/png", ".jpg": "image/jpeg",
                                 ".jpeg": "image/jpeg", ".gif": "image/gif",
                                 ".webp": "image/webp"}
                    media_type = media_map.get(ext, "image/png")
                    with open(path, "rb") as f:
                        img_b64 = base64.standard_b64encode(f.read()).decode()
                    result_content = [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": media_type,
                                "data": img_b64
                            }
                        }
                    ]
                    log_event({"event": "tool_use", "user_id": user_id, "tool": "read_image",
                               "input": block.input, "result_preview": f"画像読み込み成功: {path}", "success": True})
                except Exception as e:
                    result_content = f"画像読み込みエラー: {e}"
                    log_event({"event": "tool_use", "user_id": user_id, "tool": "read_image",
                               "input": block.input, "result_preview": result_content, "success": False})
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": block.id,
                    "content": result_content
                })

            # その他のツール
            else:
                result = await asyncio.get_event_loop().run_in_executor(
                    None, lambda b=block: execute_tool(b.name, b.input, user_id)
                )
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": block.id,
                    "content": result
                })

        messages.append({"role": "user", "content": tool_results})

        # 無限ループ防止
        if tool_count >= 20:
            return "ツール実行が20回を超えたため中断しました。"


async def send_long(message, response: str):
    """2000文字超えを分割送信"""
    if len(response) > 1900:
        chunks = [response[i:i+1900] for i in range(0, len(response), 1900)]
        for i, chunk in enumerate(chunks):
            prefix = f"[{i+1}/{len(chunks)}]\n" if len(chunks) > 1 else ""
            await message.reply(f"{prefix}```\n{chunk}\n```")
    else:
        await message.reply(response)


@client.event
async def on_ready():
    print(f"起動完了: {client.user}")
    await client.change_presence(status=discord.Status.online)


@client.event
async def on_message(message):
    if message.author == client.user:
        return
    if message.author.id not in ALLOWED_USER_IDS:
        return
    # DMのみ受け付ける
    if not isinstance(message.channel, discord.DMChannel):
        return

    content = message.content.strip()
    content_lower = content.lower()
    user_id = message.author.id

    # コマンド処理
    if content_lower == "!reset":
        history[user_id] = []
        await message.reply("🗑️ 会話履歴をリセットしました。")
        return

    if content_lower == "!help":
        await message.reply(
            "**コマンド一覧**\n"
            "`!reset` — 自分の会話履歴をリセット\n"
            "`!update` — Botを最新版に更新して再起動\n"
            "`!log [件数]` — 最近の操作ログを表示（デフォルト10件）\n"
            "`!help` — このメッセージを表示\n\n"
            "**できること**\n"
            "- ファイル読み書き・削除（ゴミ箱）\n"
            "- シェルコマンド実行\n"
            "- Excel操作\n"
            "- コード作成・修正\n"
            "- 画像解析（スクショ送付）"
        )
        return

    if content_lower.startswith("!log"):
        try:
            parts = content.split()
            n = int(parts[1]) if len(parts) > 1 else 10
            n = min(n, 50)
            if not os.path.exists(LOG_PATH):
                await message.reply("ログファイルがまだありません。")
                return
            with open(LOG_PATH, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            recent = lines[-n:]
            entries = [json.loads(l) for l in recent]
            msg = f"**最新{len(entries)}件のログ**\n"
            for e in entries:
                ts = e.get("ts", "")
                if e["event"] == "tool_use":
                    status = "✅" if e.get("success") else "❌"
                    inp_str = str(e.get("input", ""))[:60]
                    msg += f"`{ts}` {status} **{e['tool']}** `{inp_str}`\n"
                else:
                    tok = f" in:{e.get('input_tokens')} out:{e.get('output_tokens')}" if e.get('input_tokens') else ""
                    msg += f"`{ts}` 💬 {e.get('user_text', '')[:60]}{tok}\n"
            await message.reply(msg[:1900])
        except Exception as e:
            await message.reply(f"ログ取得エラー: {e}")
        return

    if content_lower == "!update":
        await message.reply("🔄 最新版を取得して再起動します...")
        try:
            script_path = os.path.abspath(__file__)
            repo_dir = os.path.dirname(script_path)
            # .gitがあればgit pull、なければraw URLからダウンロード
            if os.path.isdir(os.path.join(repo_dir, '.git')):
                result = subprocess.run(
                    ["git", "pull"], cwd=repo_dir,
                    capture_output=True, text=True, encoding='utf-8', errors='replace'
                )
                if result.returncode != 0:
                    await message.reply(f"git pull 失敗:\n```{result.stderr[:500]}```")
                    return
                detail = result.stdout.strip()[:300]
            else:
                url = "https://raw.githubusercontent.com/soma1023/discord-claude-bot/master/bot.py"
                async with aiohttp.ClientSession() as session:
                    async with session.get(url) as resp:
                        if resp.status != 200:
                            await message.reply(f"取得失敗: HTTP {resp.status}")
                            return
                        new_code = await resp.text()
                with open(script_path, 'w', encoding='utf-8') as f:
                    f.write(new_code)
                detail = "raw URLから取得"
            await message.reply(f"✅ 更新完了。再起動します。\n```{detail}```")
            # 自分以外の同名プロセスをすべて終了
            current_pid = os.getpid()
            subprocess.run(
                f'wmic process where "name=\'python.exe\' and processid!={current_pid}" '
                f'call terminate',
                shell=True, capture_output=True
            )
            subprocess.Popen([sys.executable, script_path])
            await client.close()
        except Exception as e:
            await message.reply(f"更新エラー: {e}")
        return

    has_images = any(
        a.content_type and a.content_type.startswith("image/")
        for a in message.attachments
    )

    if not content and not has_images:
        return

    async with message.channel.typing():
        try:
            if has_images:
                user_content = []
                async with aiohttp.ClientSession() as session:
                    for att in message.attachments:
                        if att.content_type and att.content_type.startswith("image/"):
                            async with session.get(att.url) as resp:
                                img_b64 = base64.standard_b64encode(await resp.read()).decode()
                                user_content.append({
                                    "type": "image",
                                    "source": {
                                        "type": "base64",
                                        "media_type": att.content_type.split(";")[0],
                                        "data": img_b64
                                    }
                                })
                user_content.append({
                    "type": "text",
                    "text": content if content else "この画像について説明してください。"
                })
            else:
                user_content = content

            response = await run_agent(user_id, user_content, message.channel)

        except Exception as e:
            response = f"エラー: {e}"

    await send_long(message, response)


client.run(TOKEN)
